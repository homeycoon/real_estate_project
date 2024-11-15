import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from dashboard.data import load_data

pd.set_option('display.max_rows', 76)
pd.set_option('display.max_colwidth', None)
pd.set_option('display.float_format', '{:.2f}'.format)


# Формируем excel-фйл, который будет выгружаться по кнопке "Выгрузить в .xlsx"
def to_excel_process():
    df = load_data()

    title = 'Общая статистика по рынку аренды коммерческой недвижимости в Казани'
    df_title = pd.DataFrame({'title': [title]})

    # subtitles
    total_ads_amount = df.id.count()
    cian_ads_amount = df.query('source == "ЦИАН"').ads_id.count()
    part_cian_ads = round(cian_ads_amount / total_ads_amount, 4)
    dom_click_ads_amount = df.query('source == "Домклик"').ads_id.count()
    part_dom_click_ads = round(dom_click_ads_amount / total_ads_amount, 4)

    df_total_ads = pd.DataFrame({'Общее\nкол-во объявлений': [total_ads_amount]})
    df_cian_ads = pd.DataFrame({'Кол-во объявлений\nс ЦИАН': [cian_ads_amount],
                                'Доля ЦИАН\nв общем\nкол-ве объявлений': [part_cian_ads]})
    df_dom_click_ads = pd.DataFrame({'Кол-во объявлений\nс Домклик': [dom_click_ads_amount],
                                     'Доля Домклик\nв общем\nкол-ве объявлений': [part_dom_click_ads]})

    '''amount'''
    cities_amount = df.query('city != "not_found"').city.nunique()
    city_areas_amount = df.query('city_area != "not_found"').city_area.nunique()
    typies_amount = df.query('building_type != "not_found"').building_type.nunique()

    '''price'''
    med_price = round(df.price.median(), 2)
    min_price = round(df.price.min(), 2)
    max_price = round(df.price.max(), 2)

    '''square'''
    med_square = round(df.square.median(), 2)
    min_square = round(df.square.min(), 2)
    max_square = round(df.square.max(), 2)

    '''price per square'''
    med_pps = round(df['price/square'].median(), 2)
    min_pps = round(df['price/square'].min(), 2)
    max_pps = round(df['price/square'].max(), 2)

    '''correlation'''
    data1 = df[['price/square', 'square']]
    correlation_pps_square = data1.corr(method='pearson')
    corr_pps_square_value = round(correlation_pps_square.iloc[0, 1], 4)

    data2 = df[['price/square', 'floor']]
    correlation_pps_floor = data2.corr(method='pearson')
    corr_pps_floor_value = round(correlation_pps_floor.iloc[0, 1], 4)

    workbook = Workbook()
    ws = workbook.active
    ws.title = 'df'

    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    with pd.ExcelWriter(excel_stream,
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists='overlay'
                        ) as writer:
        ws = writer.sheets['df']
        ws.sheet_view.showGridLines = False

        df_title.to_excel(writer, sheet_name="df", header=False, startcol=0, index=False)
        ws.row_dimensions[1].height = 25
        for i in ['C', 'E']:
            ws.column_dimensions[i].width = 45
        for j in ['B', 'D']:
            ws.column_dimensions[j].width = 23
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['F'].width = 20

        ws['A1'].alignment = Alignment(vertical='center')
        ws['A1'].font = Font(name='Aptos', size=24, bold=True)
        ws.merge_cells('A1:F1')

        ws['A3'] = 'Общее количество объявлений'
        ws['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws['A3'].font = Font(name='Aptos', size=20, bold=True)
        df_total_ads.to_excel(writer, sheet_name="df", header=False, startrow=2, startcol=1, index=False)
        ws['B3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws['B3'].font = Font(name='Aptos', size=20)
        ws['A4'] = 'в том числе:'
        ws['A4'].font = Font(name='Aptos', size=20, italic=True)
        ws['A5'] = 'ЦИАН'
        df_cian_ads.to_excel(writer, sheet_name="df", header=False, startrow=4, startcol=1, index=False)
        ws['A5'].font = Font(name='Aptos', size=20, bold=True)
        ws['B5'].font = Font(name='Aptos', size=20)
        ws['C5'].font = Font(name='Aptos', size=16)
        ws['C5'].number_format = "0.00%"
        ws['A6'] = 'ДомКлик'
        df_dom_click_ads.to_excel(writer, sheet_name="df", header=False, startrow=5, startcol=1, index=False)
        ws['A6'].font = Font(name='Aptos', size=20, bold=True)
        ws['B6'].font = Font(name='Aptos', size=20)
        ws['C6'].font = Font(name='Aptos', size=16)
        ws['C6'].number_format = "0.00%"

        ws['E3'] = 'Проанализировано:'
        ws['E3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws['E3'].font = Font(name='Aptos', size=20, bold=True)
        ws['E4'] = 'Городов'
        ws['E4'].font = Font(name='Aptos', size=20)
        ws['E5'] = 'Районов'
        ws['E5'].font = Font(name='Aptos', size=20)
        ws['E6'] = 'Типов помещений'
        ws['E6'].font = Font(name='Aptos', size=20)
        ws['F4'] = cities_amount
        ws['F4'].font = Font(name='Aptos', size=20)
        ws['F5'] = city_areas_amount
        ws['F5'].font = Font(name='Aptos', size=20)
        ws['F6'] = typies_amount
        ws['F6'].font = Font(name='Aptos', size=20)
        ws.merge_cells('E3:F3')

        ws['A8'] = 'Основные показатели'
        ws['A8'].font = Font(name='Aptos', size=20, bold=True)
        ws.merge_cells('A8:F8')
        ws['A8'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws['A10'] = 'Цена, руб.'
        ws['A10'].font = Font(name='Aptos', size=20, bold=True)
        ws['A10'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.merge_cells('A10:B10')
        ws['A11'] = 'Медианная цена'
        ws['A11'].font = Font(name='Aptos', size=20)
        ws['A12'] = 'Минимальная цена'
        ws['A12'].font = Font(name='Aptos', size=20)
        ws['A13'] = 'Максимальная цена'
        ws['A13'].font = Font(name='Aptos', size=20)

        ws['B11'] = med_price
        ws['B11'].font = Font(name='Aptos', size=20)
        ws['B11'].number_format = "0.00"
        ws['B12'] = min_price
        ws['B12'].font = Font(name='Aptos', size=20)
        ws['B13'] = max_price
        ws['B13'].font = Font(name='Aptos', size=20)

        ws['C10'] = 'Площадь, м2'
        ws['C10'].font = Font(name='Aptos', size=20, bold=True)
        ws['C10'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.merge_cells('C10:D10')
        ws['C11'] = 'Медианная площадь'
        ws['C11'].font = Font(name='Aptos', size=20)
        ws['C12'] = 'Минимальная площадь'
        ws['C12'].font = Font(name='Aptos', size=20)
        ws['C13'] = 'Максимальная площадь'
        ws['C13'].font = Font(name='Aptos', size=20)

        ws['D11'] = med_square
        ws['D11'].font = Font(name='Aptos', size=20)
        ws['D11'].number_format = "0.00"
        ws['D12'] = min_square
        ws['D12'].font = Font(name='Aptos', size=20)
        ws['D13'] = max_square
        ws['D13'].font = Font(name='Aptos', size=20)

        ws['E10'] = 'Цена за 1 м2, руб.'
        ws['E10'].font = Font(name='Aptos', size=20, bold=True)
        ws['E10'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.merge_cells('E10:F10')
        ws['E11'] = 'Медианная цена за 1 м2'
        ws['E11'].font = Font(name='Aptos', size=20)
        ws['E12'] = 'Минимальная цена за 1 м2'
        ws['E12'].font = Font(name='Aptos', size=20)
        ws['E13'] = 'Максимальная цена за 1 м2'
        ws['E13'].font = Font(name='Aptos', size=20)

        ws['F11'] = med_pps
        ws['F11'].font = Font(name='Aptos', size=20)
        ws['F12'] = min_pps
        ws['F12'].font = Font(name='Aptos', size=20)
        ws['F13'] = max_pps
        ws['F13'].font = Font(name='Aptos', size=20)

        ws['A15'] = 'Корреляция'
        ws['A15'].font = Font(name='Aptos', size=20, bold=True)
        ws.merge_cells('A15:F15')
        ws['A15'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws['A16'] = 'Цена за м2 и площадь'
        ws['A16'].font = Font(name='Aptos', size=20)
        ws['B16'] = corr_pps_square_value
        ws['B16'].font = Font(name='Aptos', size=20)
        ws['A17'] = 'Цена за м2 и этаж'
        ws['A17'].font = Font(name='Aptos', size=20)
        ws['B17'] = corr_pps_floor_value
        ws['B17'].font = Font(name='Aptos', size=20)

        no_border = Border()

        only_right_border = Border(
            right=Side(style='medium', color='000000')
        )

        # Применяем созданный стиль к диапазону ячеек
        for row in ws.iter_rows(min_row=10, max_row=13, min_col=0, max_col=6):
            for cell in row:
                if cell is row[-3] or cell is row[-5]:
                    cell.border = only_right_border

        for row in ws.iter_rows(min_row=16, max_row=20, min_col=1, max_col=7):
            for cell in row:
                cell.border = no_border

    return excel_stream
